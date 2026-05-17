import os
import shutil
import subprocess
import warnings
from copy import copy
from datetime import datetime
from pathlib import Path
from decimal import Decimal
from decimal import ROUND_FLOOR
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from sqlalchemy.orm import Session

from models import InquirySupplier, InquiryTask, InquiryTaskItem, InquiryRequest, Quotation, Supplier, LinkStatus, Contract, ContractTemplate


BASE_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = BASE_DIR / "static" / "templates"
CONTRACT_DIR = BASE_DIR / "static" / "contracts"
FONT_DIR = BASE_DIR / "static" / "fonts"
SIMSUN_PATH = FONT_DIR / "SimSun.ttf"
SYSTEM_SIMSUN_PATH = Path("C:/Windows/Fonts/simsun.ttc")
os.makedirs("static/contracts", exist_ok=True)
DEFAULT_TEMPLATE_CELLS = {
    "supplier_name": "E3",
    "buyer_name": "E4",
    "contract_no": "V3",
    "project_no": "F6",
    "total_amount_upper": "H9",
    "total_qty": "P9",
    "total_amount": "AA9",
    "sup_address": "I27",
    "sup_legal_rep": "I28",
    "sup_agent": "I29",
    "sup_phone": "I30",
    "sup_bank_name": "I31",
    "sup_bank_account": "I32",
    "sup_tax_id": "I33",
    "sup_fax": "I34",
    "sup_postal_code": "I35",
}
TEMPLATE_DYNAMIC_RULES = {
    "supplier_name": ("供方", 0, 3),
    "buyer_name": ("需方", 0, 3),
    "contract_no": ("合同号", 0, 4),
    "project_no": ("项目号", 0, 4),
    "total_amount_upper": ("合计人民币金额(大写)", 0, 6),
    "total_qty": ("数量", 2, 0),
    "total_amount": ("价税合计", 2, 0),
}


def _resolve_template_path(template_file_path: str = None) -> Path:
    candidates = []
    if template_file_path:
        custom_path = Path(template_file_path)
        if custom_path.is_absolute():
            candidates.append(custom_path)
        else:
            candidates.append(BASE_DIR / custom_path)
            candidates.append(TEMPLATE_DIR / custom_path)
    candidates.extend([
        TEMPLATE_DIR / "合同模版.xlsx",
        TEMPLATE_DIR / "合同模版.XLSX",
        BASE_DIR / "合同模版.xlsx",
        BASE_DIR / "合同模版.XLS",
    ])
    for p in candidates:
        if p.exists():
            return p
    if TEMPLATE_DIR.exists():
        fallback_files = sorted(
            [p for p in TEMPLATE_DIR.glob("*.xlsx") if p.is_file()] + [p for p in TEMPLATE_DIR.glob("*.XLSX") if p.is_file()],
            key=lambda p: p.stat().st_mtime,
            reverse=True
        )
        if fallback_files:
            return fallback_files[0]
    raise FileNotFoundError("合同模板文件不存在，请将‘合同模版.xlsx’放入 static/templates 目录")


def _get_active_contract_template(db: Session):
    return db.query(ContractTemplate).filter(ContractTemplate.is_active == True).order_by(ContractTemplate.id.desc()).first()


def _append_history_version(history_versions, pdf_path: str, event: str = "regenerated"):
    if not pdf_path:
        return history_versions or []
    versions = list(history_versions or [])
    if versions:
        last = versions[-1]
        if isinstance(last, dict) and last.get("pdf_path") == pdf_path:
            return versions
    versions.append({
        "pdf_path": pdf_path,
        "generated_at": datetime.now().isoformat(),
        "event": event,
    })
    return versions


def _normalize_text(value) -> str:
    return str(value or "").replace("：", ":").replace("\n", "").replace(" ", "").strip()


def _find_label_cell_openpyxl(ws, label: str, max_row: int = 20, max_col: int = 40):
    target = _normalize_text(label)
    row_limit = min(ws.max_row, max_row)
    col_limit = min(ws.max_column, max_col)
    for row_idx in range(1, row_limit + 1):
        for col_idx in range(1, col_limit + 1):
            current = _normalize_text(ws.cell(row_idx, col_idx).value)
            if not current:
                continue
            if target in current:
                return row_idx, col_idx
    return None


def _resolve_template_cells_for_openpyxl(ws) -> dict:
    template_cells = dict(DEFAULT_TEMPLATE_CELLS)
    for key, (label, row_offset, col_offset) in TEMPLATE_DYNAMIC_RULES.items():
        found = _find_label_cell_openpyxl(ws, label)
        if not found:
            continue
        row_idx = found[0] + row_offset
        col_idx = found[1] + col_offset
        if row_idx < 1 or col_idx < 1:
            continue
        template_cells[key] = f"{get_column_letter(col_idx)}{row_idx}"
    return template_cells


def _resolve_template_cells_for_win32(sheet) -> dict:
    template_cells = dict(DEFAULT_TEMPLATE_CELLS)
    row_limit = min(getattr(sheet.UsedRange, "Rows", sheet.UsedRange).Count, 20)
    col_limit = min(getattr(sheet.UsedRange, "Columns", sheet.UsedRange).Count, 40)
    label_positions = {}
    for row_idx in range(1, row_limit + 1):
        for col_idx in range(1, col_limit + 1):
            value = _normalize_text(sheet.Cells(row_idx, col_idx).Value)
            if not value:
                continue
            for _, (label, _, _) in TEMPLATE_DYNAMIC_RULES.items():
                if label in label_positions:
                    continue
                if _normalize_text(label) in value:
                    label_positions[label] = (row_idx, col_idx)
    for key, (label, row_offset, col_offset) in TEMPLATE_DYNAMIC_RULES.items():
        found = label_positions.get(label)
        if not found:
            continue
        row_idx = found[0] + row_offset
        col_idx = found[1] + col_offset
        if row_idx < 1 or col_idx < 1:
            continue
        template_cells[key] = f"{get_column_letter(col_idx)}{row_idx}"
    return template_cells


def _register_pdf_font() -> str:
    font_candidates = [SIMSUN_PATH, SYSTEM_SIMSUN_PATH]
    for font_path in font_candidates:
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont("SimSun", str(font_path.resolve())))
                return "SimSun"
            except Exception:
                continue
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    return "STSong-Light"


def _import_win32_modules():
    try:
        import pythoncom
        from win32com import client as win32
        return pythoncom, win32
    except Exception:
        return None, None


def _normalize_template_for_openpyxl(template_path: Path) -> Path:
    normalized_path = template_path.parent / f"_normalized_{template_path.name}"
    pythoncom, win32 = _import_win32_modules()
    if not pythoncom or not win32:
        return template_path

    excel = None
    workbook = None
    try:
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(template_path.resolve()))
        normalized_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            workbook.SaveAs(str(normalized_path.resolve()), FileFormat=51)
        except Exception:
            return template_path # 如果 SaveAs 报错被拒，直接回退返回原始路径
        return normalized_path
    except Exception as e:
        try:
            if workbook:
                workbook.Close(False)
        except Exception:
            pass
        return template_path # 如果 COM 报错，回退返回原始路径，不要把整个流程卡死
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def _safe_load_workbook(file_path: Path):
    try:
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb = load_workbook(file_path)
        has_invalid_spec_warning = any("invalid specification" in str(w.message).lower() for w in caught)
        if has_invalid_spec_warning:
            normalized_template = _normalize_template_for_openpyxl(file_path)
            return load_workbook(normalized_template)
        return wb
    except Exception:
        normalized_template = _normalize_template_for_openpyxl(file_path)
        return load_workbook(normalized_template)


def _resolve_deal_link(db: Session, inquiry_id: int) -> InquirySupplier:
    deal_link = db.query(InquirySupplier).filter(
        InquirySupplier.id == inquiry_id,
        InquirySupplier.status == LinkStatus.DEAL
    ).first()
    if deal_link:
        return deal_link

    deal_link = db.query(InquirySupplier).filter(
        InquirySupplier.task_id == inquiry_id,
        InquirySupplier.status == LinkStatus.DEAL
    ).order_by(InquirySupplier.id.desc()).first()
    if deal_link:
        return deal_link

    raise ValueError("未找到已成交的询价记录")


def _to_decimal(val) -> Decimal:
    if val is None:
        return Decimal("0")
    return Decimal(str(val))


def _format_delivery_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            pass
    text = str(value).strip()
    if not text:
        return ""
    return text[:10]


def _load_link_quotes(db: Session, link: InquirySupplier):
    quotes = db.query(Quotation).filter(
        Quotation.inquiry_supplier_id == link.id,
        Quotation.round == link.current_round
    ).all()
    if not quotes:
        quotes = db.query(Quotation).filter(
            Quotation.inquiry_supplier_id == link.id
        ).order_by(Quotation.round.desc(), Quotation.id.asc()).all()
        if quotes:
            max_round = quotes[0].round
            quotes = [q for q in quotes if q.round == max_round]
    return quotes


def _parse_link_item_allocations(link: InquirySupplier) -> dict:
    parsed = {}
    for row in (link.item_allocations or []):
        if not isinstance(row, dict):
            continue
        try:
            item_id = int(row.get("item_id"))
        except (TypeError, ValueError):
            continue
        parsed[item_id] = {
            "allocated_ratio": float(row.get("allocated_ratio")) if row.get("allocated_ratio") is not None else None,
            "allocated_qty": float(row.get("allocated_qty")) if row.get("allocated_qty") is not None else None,
        }
    return parsed


def _build_allocated_qty_map(quote_rows: list, link: InquirySupplier) -> dict:
    item_level_map = _parse_link_item_allocations(link)
    if item_level_map:
        result = {}
        for row in quote_rows:
            quote = row["quote"]
            task_item = row["task_item"]
            base_qty = row["base_qty"]
            item_cfg = item_level_map.get(int(task_item.id if task_item else quote.item_id))
            if not item_cfg:
                result[quote.id] = Decimal("0")
                continue
            if item_cfg.get("allocated_qty") is not None:
                result[quote.id] = _to_decimal(item_cfg["allocated_qty"])
                continue
            ratio = _to_decimal(item_cfg.get("allocated_ratio")) / Decimal("100")
            result[quote.id] = base_qty * ratio
        return result

    base_qty_map = {}
    total_base_qty = Decimal("0")
    for row in quote_rows:
        base_qty = row["base_qty"]
        base_qty_map[row["quote"].id] = base_qty
        total_base_qty += base_qty

    if link.allocated_qty is not None:
        allocated_total_qty = _to_decimal(link.allocated_qty)
        if len(quote_rows) <= 1:
            quote = quote_rows[0]["quote"] if quote_rows else None
            return {quote.id: allocated_total_qty} if quote else {}
        if allocated_total_qty == allocated_total_qty.to_integral_value() and total_base_qty == total_base_qty.to_integral_value() and total_base_qty > 0:
            allocated_total_int = int(allocated_total_qty)
            floors = []
            remainder_parts = []
            sum_floor = 0
            for row in quote_rows:
                exact = Decimal(str(allocated_total_int)) * row["base_qty"] / total_base_qty
                base_floor = int(exact.to_integral_value(rounding=ROUND_FLOOR))
                frac = exact - Decimal(str(base_floor))
                floors.append((row["quote"].id, base_floor))
                remainder_parts.append((frac, row["quote"].id))
                sum_floor += base_floor
            remainder = allocated_total_int - sum_floor
            remainder_parts.sort(key=lambda x: (x[0], x[1]), reverse=True)
            result = {qid: Decimal(str(floor_val)) for qid, floor_val in floors}
            for i in range(max(0, remainder)):
                qid = remainder_parts[i % len(remainder_parts)][1]
                result[qid] = result.get(qid, Decimal("0")) + Decimal("1")
            return result
        if total_base_qty > 0:
            allocated_qty_map = {}
            allocated_so_far = Decimal("0")
            for index, row in enumerate(quote_rows):
                quote_id = row["quote"].id
                if index == len(quote_rows) - 1:
                    allocated_qty_map[quote_id] = allocated_total_qty - allocated_so_far
                else:
                    qty = allocated_total_qty * row["base_qty"] / total_base_qty
                    allocated_qty_map[quote_id] = qty
                    allocated_so_far += qty
            return allocated_qty_map
        average_qty = allocated_total_qty / Decimal(str(len(quote_rows)))
        return {row["quote"].id: average_qty for row in quote_rows}

    if link.allocated_ratio is not None:
        ratio = _to_decimal(link.allocated_ratio) / Decimal("100")
        return {quote_id: base_qty * ratio for quote_id, base_qty in base_qty_map.items()}

    return base_qty_map


def _build_task_split_allocated_qty_map(db: Session, quote_rows: list, link: InquirySupplier) -> dict:
    """
    拆单定标时，将每个物料的总数量按各成交供应商的 allocated_ratio 分配为整数，且所有合同数量之和等于总量。
    使用最大余数法分配尾差（四舍五入效果更接近业务预期，但保证总和一致）。
    """
    deal_links = (
        db.query(InquirySupplier)
        .filter(InquirySupplier.task_id == link.task_id, InquirySupplier.status == LinkStatus.DEAL)
        .order_by(InquirySupplier.id.asc())
        .all()
    )
    if not deal_links or len(deal_links) <= 1:
        return {}

    parsed_by_link_id = {deal_link.id: _parse_link_item_allocations(deal_link) for deal_link in deal_links}
    if any(parsed_by_link_id.values()):
        result_for_current = {}
        for row in quote_rows:
            quote = row["quote"]
            task_item = row["task_item"]
            item_id = int(task_item.id if task_item else quote.item_id)
            item_level_configs = {}
            has_item_level_config = False
            has_explicit_qty = False

            for deal_link in deal_links:
                cfg = parsed_by_link_id.get(deal_link.id, {}).get(item_id)
                if not cfg:
                    continue
                has_item_level_config = True
                if cfg.get("allocated_qty") is not None:
                    has_explicit_qty = True
                item_level_configs[deal_link.id] = cfg

            if not has_item_level_config:
                continue

            if has_explicit_qty:
                result_for_current[quote.id] = _to_decimal(item_level_configs.get(link.id, {}).get("allocated_qty"))
                continue

            current_ratio = _to_decimal(item_level_configs.get(link.id, {}).get("allocated_ratio")) / Decimal("100")
            base_qty = row["base_qty"]
            if current_ratio <= 0:
                result_for_current[quote.id] = Decimal("0")
                continue
            if base_qty != base_qty.to_integral_value():
                result_for_current[quote.id] = base_qty * current_ratio
                continue

            total_int = int(base_qty)
            floors = {}
            remainders = []
            sum_floor = 0
            for deal_link in deal_links:
                ratio = _to_decimal(item_level_configs.get(deal_link.id, {}).get("allocated_ratio")) / Decimal("100")
                exact = Decimal(str(total_int)) * ratio
                floor_val = int(exact.to_integral_value(rounding=ROUND_FLOOR))
                frac = exact - Decimal(str(floor_val))
                floors[deal_link.id] = floor_val
                remainders.append((frac, deal_link.id))
                sum_floor += floor_val

            remainder = total_int - sum_floor
            remainders.sort(key=lambda x: (x[0], x[1]), reverse=True)
            for i in range(max(0, remainder)):
                l_id = remainders[i % len(remainders)][1]
                floors[l_id] = floors.get(l_id, 0) + 1

            result_for_current[quote.id] = Decimal(str(floors.get(link.id, 0)))

        if result_for_current:
            return result_for_current

    ratios = {}
    for l in deal_links:
        if l.allocated_ratio is None:
            ratios[l.id] = Decimal("0")
        else:
            ratios[l.id] = _to_decimal(l.allocated_ratio) / Decimal("100")

    result_for_current = {}
    for row in quote_rows:
        qid = row["quote"].id
        base_qty = row["base_qty"]
        if base_qty != base_qty.to_integral_value():
            continue

        total_int = int(base_qty)
        floors = {}
        remainders = []
        sum_floor = 0
        for l in deal_links:
            exact = Decimal(str(total_int)) * ratios.get(l.id, Decimal("0"))
            floor_val = int(exact.to_integral_value(rounding=ROUND_FLOOR))
            frac = exact - Decimal(str(floor_val))
            floors[l.id] = floor_val
            remainders.append((frac, l.id))
            sum_floor += floor_val

        remainder = total_int - sum_floor
        remainders.sort(key=lambda x: (x[0], x[1]), reverse=True)
        for i in range(max(0, remainder)):
            l_id = remainders[i % len(remainders)][1]
            floors[l_id] = floors.get(l_id, 0) + 1

        current_alloc = floors.get(link.id)
        if current_alloc is not None:
            result_for_current[qid] = Decimal(str(current_alloc))

    return result_for_current


def _estimate_wrapped_lines(value, chars_per_line: int) -> int:
    if chars_per_line <= 0:
        chars_per_line = 1
    text = str(value or "")
    if not text:
        return 1
    total_lines = 0
    for segment in text.replace("\r\n", "\n").split("\n"):
        seg_len = len(segment)
        if seg_len == 0:
            total_lines += 1
        else:
            total_lines += (seg_len + chars_per_line - 1) // chars_per_line
    return max(1, total_lines)


def _to_chinese_upper_amount(amount: Decimal) -> str:
    digits = "零壹贰叁肆伍陆柒捌玖"
    units = ["", "拾", "佰", "仟"]
    big_units = ["", "万", "亿", "兆"]
    integer = int(amount.quantize(Decimal("1")))
    if integer == 0:
        return "零圆整"
    groups = []
    while integer > 0:
        groups.append(integer % 10000)
        integer //= 10000
    text_parts = []
    for gi in range(len(groups) - 1, -1, -1):
        group = groups[gi]
        if group == 0:
            if text_parts and not text_parts[-1].endswith("零"):
                text_parts.append("零")
            continue
        group_text = ""
        zero_flag = False
        for pos in range(3, -1, -1):
            divisor = 10 ** pos
            n = group // divisor
            group %= divisor
            if n == 0:
                zero_flag = True
            else:
                if zero_flag and group_text and not group_text.endswith("零"):
                    group_text += "零"
                zero_flag = False
                group_text += digits[n] + units[pos]
        text_parts.append(group_text + big_units[gi])
    return f"{''.join(text_parts).rstrip('零')}圆整"


def _collect_contract_payload(
    db: Session,
    link: InquirySupplier,
    contract_record: Contract = None,
    buyer_company_name: str = None
) -> dict:
    task = db.query(InquiryTask).filter(InquiryTask.id == link.task_id).first()
    supplier = db.query(Supplier).filter(Supplier.id == link.supplier_id).first()
    if not task or not supplier:
        raise ValueError("询价任务或供应商信息不存在")

    quotes = _load_link_quotes(db, link)
    if not quotes:
        raise ValueError("未找到该成交供应商的报价数据")

    project_no = ""
    project_name = ""
    buyer_name = buyer_company_name or "俊朗电气有限公司"
    if contract_record and contract_record.buyer_company_name:
        buyer_name = contract_record.buyer_company_name
    items = []
    total_amount = Decimal("0")
    total_qty = Decimal("0")
    quote_rows = []
    for q in quotes:
        task_item = db.query(InquiryTaskItem).filter(InquiryTaskItem.id == q.item_id).first()
        req = db.query(InquiryRequest).filter(InquiryRequest.id == task_item.request_id).first() if task_item else None
        base_qty = _to_decimal(req.qty if req and req.qty is not None else q.qty)
        quote_rows.append({
            "quote": q,
            "task_item": task_item,
            "request": req,
            "base_qty": base_qty,
        })
    allocated_qty_map = _build_task_split_allocated_qty_map(db, quote_rows, link)
    if not allocated_qty_map:
        allocated_qty_map = _build_allocated_qty_map(quote_rows, link)

    for idx, row in enumerate(quote_rows, start=1):
        q = row["quote"]
        req = row["request"]
        material_name = req.material_name if req else ""
        material_code = req.material_code if req else ""
        item_project_no = str(req.project_info.get("number") or req.project_info.get("name") or "") if req and req.project_info else ""
        item_project_name = str(req.project_info.get("name") or req.project_info.get("number") or "") if req and req.project_info else ""
        qty = allocated_qty_map.get(q.id, row["base_qty"])
        if qty <= 0:
            continue
        price = _to_decimal(q.price)
        amount = qty * price
        total_qty += qty
        total_amount += amount

        if not project_no and item_project_no:
            project_no = item_project_no
        if not project_name and item_project_name:
            project_name = item_project_name

        items.append({
            "index": len(items) + 1,
            "project_no": item_project_no,
            "project_name": item_project_name,
            "material_name": material_name,
            "material_code": material_code,
            "qty": int(qty) if qty == qty.to_integral_value() else float(qty),
            "price": float(price),
            "amount": float(amount),
            "delivery_date": _format_delivery_date(q.delivery_date or (req.delivery_date if req else None)),
        })

    return {
        "contract_no": f"HT-{link.task_id}-{link.id}-{datetime.now().strftime('%Y%m%d')}",
        "supplier_name": supplier.name,
        "buyer_name": buyer_name,
        "project_no": project_no,
        "project_name": project_name,
        "task_title": task.title,
        "items": items,
        "total_qty": int(total_qty) if total_qty == total_qty.to_integral_value() else float(total_qty),
        "total_amount": float(total_amount),
        "sup_address": contract_record.address if contract_record else "",
        "sup_legal_rep": contract_record.legal_representative if contract_record else "",
        "sup_agent": contract_record.agent if contract_record else "",
        "sup_phone": contract_record.contact_phone if contract_record else "",
        "sup_bank_name": contract_record.bank_name if contract_record else "",
        "sup_bank_account": contract_record.bank_account if contract_record else "",
        "sup_tax_id": contract_record.tax_id if contract_record else "",
        "sup_fax": contract_record.fax if contract_record else "",
        "sup_postal_code": contract_record.postal_code if contract_record else "",
    }


def _fill_template_excel(payload: dict, output_xlsx: Path, template_path: Path = None) -> None:
    template_path = template_path or _resolve_template_path()
    wb = _safe_load_workbook(template_path)
    if not wb.worksheets:
        normalized_template = _normalize_template_for_openpyxl(template_path)
        wb = load_workbook(normalized_template)
    if not wb.worksheets:
        raise ValueError("模板文件不包含工作表")
    ws = wb.active if wb.active else wb.worksheets[0]
    template_cells = _resolve_template_cells_for_openpyxl(ws)

    def resolve_cell_ref(cell_ref: str) -> str:
        try:
            cell = ws[cell_ref]
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        return ws.cell(row=merged_range.min_row, column=merged_range.min_col).coordinate
            return cell_ref
        except Exception:
            return cell_ref

    def set_cell_value(cell_ref: str, value):
        try:
            target_ref = resolve_cell_ref(cell_ref)
            ws[target_ref] = value
        except Exception:
            return

    def set_item_cell_value(row_idx: int, col_idx: int, value):
        try:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                        return
            ws.cell(row=row_idx, column=col_idx).value = value
        except Exception:
            return

    def get_item_cell(row_idx: int, col_idx: int):
        try:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return cell
        except Exception:
            return ws.cell(row=row_idx, column=col_idx)

    def clone_row_style(source_row: int, target_row: int):
        try:
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
            for col in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=source_row, column=col)
                target_cell = ws.cell(row=target_row, column=col)
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
                    target_cell.protection = copy(source_cell.protection)
                    target_cell.alignment = copy(source_cell.alignment)
        except Exception:
            return

    set_cell_value(template_cells["supplier_name"], payload.get("supplier_name", ""))
    set_cell_value(template_cells["buyer_name"], payload.get("buyer_name", ""))
    set_cell_value(template_cells["contract_no"], payload.get("contract_no", ""))
    set_cell_value(template_cells["project_no"], payload.get("project_no") or payload.get("task_title", ""))
    items = payload.get("items", [])
    row = 42
    base_item_row_height = ws.row_dimensions[row].height or 22
    max_item_row = ws.max_row
    remark_row = None
    for r in range(42, ws.max_row + 1):
        marker = ws.cell(row=r, column=2).value
        if isinstance(marker, str) and "备注" in marker:
            remark_row = r
            max_item_row = r - 1
            break
    if items and remark_row is not None:
        current_capacity = max_item_row - row + 1
        if current_capacity < 0:
            current_capacity = 0
        remark_row_height = ws.row_dimensions[remark_row].height
        remark_row_values = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=remark_row, column=col).value
            if value not in (None, ""):
                remark_row_values[col] = value
        remark_row_merges = []
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == remark_row and merged_range.max_row == remark_row:
                remark_row_merges.append((merged_range.min_col, merged_range.max_col))
        base_row_merges = []
        if current_capacity > 0:
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row == row and merged_range.max_row == row:
                    base_row_merges.append((merged_range.min_col, merged_range.max_col))

        def reset_item_row_structure(target_row: int):
            clone_row_style(row, target_row)
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == target_row and merged_range.max_row == target_row:
                    ws.unmerge_cells(str(merged_range))
            for min_col, max_col in base_row_merges:
                ws.merge_cells(
                    start_row=target_row,
                    start_column=min_col,
                    end_row=target_row,
                    end_column=max_col
                )

        def reset_remark_row_structure(target_row: int):
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row == target_row and merged_range.max_row == target_row:
                    ws.unmerge_cells(str(merged_range))
            ws.row_dimensions[target_row].height = remark_row_height
            for min_col, max_col in remark_row_merges:
                ws.merge_cells(
                    start_row=target_row,
                    start_column=min_col,
                    end_row=target_row,
                    end_column=max_col
                )
            for col, value in remark_row_values.items():
                set_item_cell_value(target_row, col, value)

        final_remark_row = remark_row
        if len(items) > current_capacity:
            extra_rows = len(items) - current_capacity
            ws.insert_rows(remark_row, amount=extra_rows)
            for insert_idx in range(extra_rows):
                target_row = remark_row + insert_idx
                reset_item_row_structure(target_row)
            max_item_row = remark_row + extra_rows - 1
            final_remark_row = remark_row + extra_rows
        reset_remark_row_structure(final_remark_row)
        for target_row in range(row + 1, row + len(items)):
            if target_row <= max_item_row:
                reset_item_row_structure(target_row)
    for item in items:
        if row > max_item_row:
            break
        item_project_no = item.get("project_no") or payload.get("project_no") or payload.get("task_title", "")
        item_project_name = item.get("project_name") or payload.get("project_name") or payload.get("task_title", "")
        item_material_name = item.get("material_name", "")
        item_material_code = item.get("material_code", "")
        item_delivery_date = item.get("delivery_date", "")
        item_cols = {
            2: item.get("index"),
            3: item_project_no,
            7: item_project_name,
            9: item_material_name,
            11: item_material_code,
            15: item.get("qty", 0),
            19: item.get("price", 0),
            26: item.get("amount", 0),
            29: item_delivery_date,
        }
        base_row = 42
        for col_idx, value in item_cols.items():
            set_item_cell_value(row, col_idx, value)
            if row != base_row:
                source_cell = ws.cell(row=base_row, column=col_idx)
                target_cell = get_item_cell(row, col_idx)
                if source_cell.has_style:
                    target_cell.font = copy(source_cell.font)
                    target_cell.border = copy(source_cell.border)
                    target_cell.fill = copy(source_cell.fill)
                    target_cell.number_format = source_cell.number_format
        wrap_rules = {
            3: (item_project_no, 12),
            7: (item_project_name, 7),
            9: (item_material_name, 8),
            11: (item_material_code, 14),
            29: (item_delivery_date, 10),
        }
        max_lines = 1
        for col_idx, (text, chars_per_line) in wrap_rules.items():
            target_cell = get_item_cell(row, col_idx)
            base_cell = ws.cell(row=base_row, column=col_idx)
            alignment = copy(base_cell.alignment) if base_cell.alignment else copy(target_cell.alignment)
            alignment.wrap_text = True
            target_cell.alignment = alignment
            if row != base_row and base_cell.has_style:
                target_cell.font = copy(base_cell.font)
            max_lines = max(max_lines, _estimate_wrapped_lines(text, chars_per_line))
        ws.row_dimensions[row].height = max(base_item_row_height, 18 * max_lines + 10)
        row += 1
    total_amount = _to_decimal(payload.get("total_amount", 0))
    total_qty = _to_decimal(payload.get("total_qty", 0))
    set_cell_value(template_cells["total_amount_upper"], _to_chinese_upper_amount(total_amount))
    set_cell_value(template_cells["total_qty"], float(total_qty))
    set_cell_value(template_cells["total_amount"], float(total_amount))
    set_cell_value(template_cells["sup_address"], payload.get("sup_address", ""))
    set_cell_value(template_cells["sup_legal_rep"], payload.get("sup_legal_rep", ""))
    set_cell_value(template_cells["sup_agent"], payload.get("sup_agent", ""))
    set_cell_value(template_cells["sup_phone"], payload.get("sup_phone", ""))
    set_cell_value(template_cells["sup_bank_name"], payload.get("sup_bank_name", ""))
    set_cell_value(template_cells["sup_bank_account"], payload.get("sup_bank_account", ""))
    set_cell_value(template_cells["sup_tax_id"], payload.get("sup_tax_id", ""))
    set_cell_value(template_cells["sup_fax"], payload.get("sup_fax", ""))
    set_cell_value(template_cells["sup_postal_code"], payload.get("sup_postal_code", ""))

    base_item_row = 42
    item_data_cols = [2, 3, 7, 9, 11, 15, 19, 26, 29]
    for check_row in range(base_item_row + 1, row):
        if check_row > max_item_row:
            break
        for col_idx in item_data_cols:
            try:
                source_cell = ws.cell(row=base_item_row, column=col_idx)
                target_cell = ws.cell(row=check_row, column=col_idx)
                if source_cell.has_style and source_cell.font and source_cell.font.name:
                    from openpyxl.styles import Font
                    target_cell.font = Font(
                        name=source_cell.font.name,
                        size=source_cell.font.size,
                        bold=source_cell.font.bold,
                        italic=source_cell.font.italic,
                        color=source_cell.font.color,
                    )
            except Exception:
                pass

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def _fill_template_excel_with_win32(payload: dict, output_xlsx: Path, template_path: Path = None) -> bool:
    pythoncom, win32 = _import_win32_modules()
    if not pythoncom or not win32:
        return False
    template_path = template_path or _resolve_template_path()
    excel = None
    workbook = None
    try:
        import pywintypes
    except Exception:
        pywintypes = None

    def _call_with_retry(func, retries: int = 8, wait_seconds: float = 0.4):
        import time
        last_error = None
        for _ in range(retries):
            try:
                return func()
            except Exception as exc:
                last_error = exc
                if pywintypes is not None and isinstance(exc, pywintypes.com_error):
                    if exc.args and exc.args[0] == -2147418111:
                        try:
                            pythoncom.PumpWaitingMessages()
                        except Exception:
                            pass
                        time.sleep(wait_seconds)
                        continue
                raise
        if last_error:
            raise last_error

    try:
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = _call_with_retry(lambda: excel.Workbooks.Open(str(template_path.resolve())))
        sheet = workbook.Worksheets(1)
        template_cells = _resolve_template_cells_for_win32(sheet)

        def set_range_value(cell_ref: str, value):
            try:
                cell_range = sheet.Range(cell_ref)
                if bool(cell_range.MergeCells):
                    cell_range.MergeArea.Cells(1, 1).Value = value
                else:
                    cell_range.Value = value
            except Exception:
                return

        def set_item_cell_value(row_idx: int, col_idx: int, value):
            try:
                cell = sheet.Cells(row_idx, col_idx)
                if bool(cell.MergeCells):
                    cell.MergeArea.Cells(1, 1).Value = value
                else:
                    cell.Value = value
            except Exception:
                return

        set_range_value(template_cells["supplier_name"], payload.get("supplier_name", ""))
        set_range_value(template_cells["buyer_name"], payload.get("buyer_name", ""))
        set_range_value(template_cells["contract_no"], payload.get("contract_no", ""))
        set_range_value(template_cells["project_no"], payload.get("project_no") or payload.get("task_title", ""))
        items = payload.get("items", [])
        row = 42
        base_row_num = row
        base_item_row_height = float(sheet.Rows(row).RowHeight or 22)
        max_item_row = int(sheet.UsedRange.Rows.Count)
        remark_row = None
        for r in range(42, max_item_row + 1):
            marker = sheet.Cells(r, 2).Value
            if isinstance(marker, str) and "备注" in marker:
                remark_row = r
                max_item_row = r - 1
                break

        def _collect_single_row_merges(target_row: int):
            merges = []
            seen = set()
            used_cols = int(sheet.UsedRange.Columns.Count)
            for col_idx in range(1, used_cols + 1):
                cell = sheet.Cells(target_row, col_idx)
                if not bool(cell.MergeCells):
                    continue
                area = cell.MergeArea
                if int(area.Row) != target_row or int(area.Rows.Count) != 1:
                    continue
                start_col = int(area.Column)
                end_col = start_col + int(area.Columns.Count) - 1
                key = (start_col, end_col)
                if key in seen:
                    continue
                seen.add(key)
                merges.append(key)
            return merges

        base_row_merges = _collect_single_row_merges(base_row_num)

        def _reset_item_row_structure(target_row: int):
            used_cols = int(sheet.UsedRange.Columns.Count)
            cleared_addresses = set()
            for col_idx in range(1, used_cols + 1):
                cell = sheet.Cells(target_row, col_idx)
                if not bool(cell.MergeCells):
                    continue
                area = cell.MergeArea
                if int(area.Row) != target_row or int(area.Rows.Count) != 1:
                    continue
                area_address = str(area.Address)
                if area_address in cleared_addresses:
                    continue
                cleared_addresses.add(area_address)
                area.UnMerge()
            for start_col, end_col in base_row_merges:
                if start_col == end_col:
                    continue
                sheet.Range(
                    sheet.Cells(target_row, start_col),
                    sheet.Cells(target_row, end_col)
                ).Merge()

        if items and remark_row is not None:
            current_capacity = max_item_row - row + 1
            if current_capacity < 0:
                current_capacity = 0
            if len(items) > current_capacity:
                extra_rows = len(items) - current_capacity
                sheet.Rows(f"{remark_row}:{remark_row + extra_rows - 1}").Insert()
                for insert_idx in range(extra_rows):
                    _reset_item_row_structure(remark_row + insert_idx)
                max_item_row = remark_row + extra_rows - 1
        for item in items:
            if row > max_item_row:
                break
            item_project_no = item.get("project_no") or payload.get("project_no") or payload.get("task_title", "")
            item_project_name = item.get("project_name") or payload.get("project_name") or payload.get("task_title", "")
            item_material_name = item.get("material_name", "")
            item_material_code = item.get("material_code", "")
            item_delivery_date = item.get("delivery_date", "")
            set_item_cell_value(row, 2, item.get("index"))
            set_item_cell_value(row, 3, item_project_no)
            set_item_cell_value(row, 7, item_project_name)
            set_item_cell_value(row, 9, item_material_name)
            set_item_cell_value(row, 11, item_material_code)
            set_item_cell_value(row, 15, item.get("qty", 0))
            set_item_cell_value(row, 19, item.get("price", 0))
            set_item_cell_value(row, 26, item.get("amount", 0))
            set_item_cell_value(row, 29, item_delivery_date)
            if row != base_row_num:
                try:
                    for col_idx in [2, 3, 7, 9, 11, 15, 19, 26, 29]:
                        source_cell = sheet.Cells(base_row_num, col_idx)
                        target_cell = sheet.Cells(row, col_idx)
                        target_cell.Font.Name = source_cell.Font.Name
                        target_cell.Font.Size = source_cell.Font.Size
                except Exception:
                    pass
            wrap_rules = {
                3: (item_project_no, 12),
                7: (item_project_name, 7),
                9: (item_material_name, 8),
                11: (item_material_code, 14),
                29: (item_delivery_date, 10),
            }
            max_lines = 1
            for col_idx, (text, chars_per_line) in wrap_rules.items():
                cell = sheet.Cells(row, col_idx)
                target = cell.MergeArea if bool(cell.MergeCells) else cell
                target.WrapText = True
                max_lines = max(max_lines, _estimate_wrapped_lines(text, chars_per_line))
            sheet.Rows(row).RowHeight = max(base_item_row_height, 18 * max_lines + 10)
            row += 1
        total_amount = _to_decimal(payload.get("total_amount", 0))
        total_qty = _to_decimal(payload.get("total_qty", 0))
        set_range_value(template_cells["total_amount_upper"], _to_chinese_upper_amount(total_amount))
        set_range_value(template_cells["total_qty"], float(total_qty))
        set_range_value(template_cells["total_amount"], float(total_amount))
        set_range_value(template_cells["sup_address"], payload.get("sup_address", ""))
        set_range_value(template_cells["sup_legal_rep"], payload.get("sup_legal_rep", ""))
        set_range_value(template_cells["sup_agent"], payload.get("sup_agent", ""))
        set_range_value(template_cells["sup_phone"], payload.get("sup_phone", ""))
        set_range_value(template_cells["sup_bank_name"], payload.get("sup_bank_name", ""))
        set_range_value(template_cells["sup_bank_account"], payload.get("sup_bank_account", ""))
        set_range_value(template_cells["sup_tax_id"], payload.get("sup_tax_id", ""))
        set_range_value(template_cells["sup_fax"], payload.get("sup_fax", ""))
        set_range_value(template_cells["sup_postal_code"], payload.get("sup_postal_code", ""))
        try:
            base_row_num = 42
            font_cols = [2, 3, 7, 9, 11, 15, 19, 26, 29]
            for check_row in range(base_row_num + 1, row):
                for col_idx in font_cols:
                    try:
                        src = sheet.Cells(base_row_num, col_idx)
                        tgt = sheet.Cells(check_row, col_idx)
                        if src.Font.Name:
                            tgt.Font.Name = src.Font.Name
                            tgt.Font.Size = src.Font.Size
                            tgt.Font.Bold = src.Font.Bold
                    except Exception:
                        pass
        except Exception:
            pass
        output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        _call_with_retry(lambda: workbook.SaveAs(str(output_xlsx.resolve()), FileFormat=51))
        return True
    except Exception:
        return False
    finally:
        if workbook is not None:
            try:
                _call_with_retry(lambda: workbook.Close(False), retries=4, wait_seconds=0.2)
            except Exception:
                pass
        if excel is not None:
            try:
                _call_with_retry(lambda: excel.Quit(), retries=4, wait_seconds=0.2)
            except Exception:
                pass
        if pythoncom is not None:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def _fill_template_to_temp_excel(payload: dict, template_path: Path = None) -> Path:
    import logging
    logger = logging.getLogger(__name__)
    temp_xlsx = CONTRACT_DIR / f"temp_filled_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:8]}.xlsx"
    template_path = template_path or _resolve_template_path()
    if _fill_template_excel_with_win32(payload, temp_xlsx, template_path=template_path):
        logger.info(f"Excel填充成功: 使用Win32(Excel)COM引擎 -> {temp_xlsx.name}")
        return temp_xlsx
    logger.info(f"Excel填充: Win32失败,使用openpyxl引擎 -> {temp_xlsx.name}")
    _fill_template_excel(payload, temp_xlsx, template_path=template_path)
    return temp_xlsx


def _export_excel_to_pdf_with_win32(xlsx_path: Path, output_pdf: Path) -> bool:
    pythoncom, win32 = _import_win32_modules()
    if not pythoncom or not win32:
        return False

    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    try:
        import pywintypes
    except Exception:
        pywintypes = None

    def _call_with_retry(func, retries: int = 12, wait_seconds: float = 0.5):
        import time
        last_error = None
        for _ in range(retries):
            try:
                return func()
            except Exception as exc:
                last_error = exc
                if pywintypes is not None and isinstance(exc, pywintypes.com_error):
                    if exc.args and exc.args[0] == -2147418111:
                        try:
                            pythoncom.PumpWaitingMessages()
                        except Exception:
                            pass
                        time.sleep(wait_seconds)
                        continue
                raise
        if last_error:
            raise last_error

    # Recreate Excel application for each attempt to avoid poisoned COM state.
    for _ in range(3):
        excel = None
        workbook = None
        try:
            pythoncom.CoInitialize()
            excel = win32.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = _call_with_retry(lambda: excel.Workbooks.Open(str(xlsx_path.resolve())))
            _call_with_retry(lambda: workbook.ExportAsFixedFormat(0, str(output_pdf.resolve())))
            return True
        except Exception:
            pass
        finally:
            if workbook is not None:
                try:
                    _call_with_retry(lambda: workbook.Close(False), retries=4, wait_seconds=0.2)
                except Exception:
                    pass
            if excel is not None:
                try:
                    _call_with_retry(lambda: excel.Quit(), retries=4, wait_seconds=0.2)
                except Exception:
                    pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
    return False


def _export_excel_to_pdf_with_wps(xlsx_path: Path, output_pdf: Path) -> bool:
    pythoncom, win32 = _import_win32_modules()
    if not pythoncom or not win32:
        return False

    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    try:
        import pywintypes
    except Exception:
        pywintypes = None

    def _call_with_retry(func, retries: int = 10, wait_seconds: float = 0.4):
        import time
        last_error = None
        for _ in range(retries):
            try:
                return func()
            except Exception as exc:
                last_error = exc
                if pywintypes is not None and isinstance(exc, pywintypes.com_error):
                    if exc.args and exc.args[0] == -2147418111:
                        try:
                            pythoncom.PumpWaitingMessages()
                        except Exception:
                            pass
                        time.sleep(wait_seconds)
                        continue
                raise
        if last_error:
            raise last_error

    # WPS spreadsheets COM ProgID commonly uses ket.Application or et.Application.
    for prog_id in ("ket.Application", "et.Application"):
        for _ in range(2):
            app = None
            workbook = None
            try:
                pythoncom.CoInitialize()
                app = win32.DispatchEx(prog_id)
                app.Visible = False
                app.DisplayAlerts = False
                workbook = _call_with_retry(lambda: app.Workbooks.Open(str(xlsx_path.resolve())))
                try:
                    _call_with_retry(lambda: workbook.ExportAsFixedFormat(0, str(output_pdf.resolve())))
                except Exception:
                    # Some WPS versions only support SaveAs(FileFormat=57) for PDF.
                    _call_with_retry(lambda: workbook.SaveAs(str(output_pdf.resolve()), 57))
                return output_pdf.exists() and output_pdf.stat().st_size > 0
            except Exception:
                pass
            finally:
                if workbook is not None:
                    try:
                        _call_with_retry(lambda: workbook.Close(False), retries=4, wait_seconds=0.2)
                    except Exception:
                        pass
                if app is not None:
                    try:
                        _call_with_retry(lambda: app.Quit(), retries=4, wait_seconds=0.2)
                    except Exception:
                        pass
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    return False


def _find_soffice_executable() -> Path | None:
    env_candidates = [
        os.getenv("SOFFICE_PATH"),
        os.getenv("LIBREOFFICE_PATH"),
    ]
    for candidate in env_candidates:
        if candidate and Path(candidate).exists():
            return Path(candidate)

    for cmd_name in ("soffice", "libreoffice"):
        resolved = shutil.which(cmd_name)
        if resolved and Path(resolved).exists():
            return Path(resolved)

    program_files = [
        os.getenv("ProgramFiles"),
        os.getenv("ProgramFiles(x86)"),
    ]
    for base in program_files:
        if not base:
            continue
        p = Path(base) / "LibreOffice" / "program" / "soffice.exe"
        if p.exists():
            return p
    return None


def _export_excel_to_pdf_with_libreoffice(xlsx_path: Path, output_pdf: Path) -> bool:
    soffice_path = _find_soffice_executable()
    if soffice_path is None:
        return False

    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    generated_pdf = output_pdf.parent / f"{xlsx_path.stem}.pdf"
    if generated_pdf.exists():
        try:
            generated_pdf.unlink()
        except Exception:
            pass

    command = [
        str(soffice_path),
        "--headless",
        "--nologo",
        "--nodefault",
        "--norestore",
        "--nolockcheck",
        "--convert-to",
        "pdf",
        "--outdir",
        str(output_pdf.parent.resolve()),
        str(xlsx_path.resolve()),
    ]
    try:
        result = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=120,
        )
    except Exception:
        return False

    if result.returncode != 0:
        return False
    if not generated_pdf.exists():
        return False

    if generated_pdf.resolve() != output_pdf.resolve():
        try:
            if output_pdf.exists():
                output_pdf.unlink()
        except Exception:
            pass
        generated_pdf.replace(output_pdf)

    return output_pdf.exists() and output_pdf.stat().st_size > 0


def _export_excel_to_pdf(xlsx_path: Path, output_pdf: Path, payload: dict = None) -> None:
    import logging
    logger = logging.getLogger(__name__)
    if _export_excel_to_pdf_with_wps(xlsx_path, output_pdf):
        logger.info(f"PDF导出成功: 使用WPS引擎 -> {output_pdf.name}")
        return
    if _export_excel_to_pdf_with_libreoffice(xlsx_path, output_pdf):
        logger.info(f"PDF导出成功: 使用LibreOffice引擎 -> {output_pdf.name}")
        return
    if _export_excel_to_pdf_with_win32(xlsx_path, output_pdf):
        logger.info(f"PDF导出成功: 使用Win32(Excel)COM引擎 -> {output_pdf.name}")
        return
    logger.warning(f"PDF导出: 所有COM引擎失败,使用reportlab兜底渲染 -> {output_pdf.name}")
    _render_pdf_with_reportlab(xlsx_path, output_pdf, payload=payload)


def _render_pdf_with_reportlab(xlsx_path: Path, output_pdf: Path, payload: dict = None) -> None:
    wb = _safe_load_workbook(xlsx_path)
    ws = wb.active if wb.active else wb.worksheets[0]
    template_cells = _resolve_template_cells_for_openpyxl(ws)
    font_name = _register_pdf_font()

    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(output_pdf.resolve()))
    c.setPageSize((595, 842))
    c.setFont(font_name, 11)

    start_x = 40
    y = 800
    base_line_spacing = 20

    c.setFont(font_name, 16)
    c.drawString(start_x + 210, y, "采购合同")
    y -= 32
    c.setFont(font_name, 11)

    def safe_str(value):
        if value is None:
            return ""
        s = str(value)
        try:
            s.encode('utf-8')
            return s
        except (UnicodeEncodeError, UnicodeDecodeError):
            return "".join(c for c in s if ord(c) < 127 or ord(c) > 0x4e00)

    def get_display_value_by_ref(cell_ref: str):
        try:
            cell = ws[cell_ref]
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return cell.value
        except Exception:
            return ""

    def get_display_value_by_pos(row_idx: int, col_idx: int):
        try:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return cell.value
        except Exception:
            return ""

    header_refs = [
        ("供方", template_cells["supplier_name"]),
        ("需方", template_cells["buyer_name"]),
        ("合同号", template_cells["contract_no"]),
        ("项目号", template_cells["project_no"]),
    ]
    for label, ref in header_refs:
        val = safe_str(get_display_value_by_ref(ref))
        c.setFont(font_name, 11)
        c.drawString(start_x, y, f"{label}：{val}")
        y -= base_line_spacing

    supplier_extra_refs = [
        ("供方地址", template_cells["sup_address"]),
        ("法定代表人", template_cells["sup_legal_rep"]),
        ("委托代理人", template_cells["sup_agent"]),
        ("联系电话", template_cells["sup_phone"]),
        ("开户银行", template_cells["sup_bank_name"]),
        ("账号", template_cells["sup_bank_account"]),
        ("税号", template_cells["sup_tax_id"]),
        ("传真", template_cells["sup_fax"]),
        ("邮编", template_cells["sup_postal_code"]),
    ]
    for label, ref in supplier_extra_refs:
        val = safe_str(get_display_value_by_ref(ref))
        c.setFont(font_name, 10)
        c.drawString(start_x, y, f"{label}：{val}")
        y -= 16

    y -= 10
    table_headers = ["序号", "项目号", "项目名称", "物料名称", "型号规格", "数量", "含税单价", "价税合计", "交货日期"]
    col_widths = [32, 62, 75, 95, 95, 48, 68, 70, 70]
    x = start_x
    for i, h in enumerate(table_headers):
        c.setFont(font_name, 11)
        c.drawString(x, y, str(h))
        x += col_widths[i]
    y -= base_line_spacing

    if payload and payload.get("items"):
        items = payload["items"]
        project_no = payload.get("project_no") or payload.get("task_title", "")
        project_name = payload.get("project_name") or payload.get("task_title", "")
        for item in items:
            if y <= 80:
                break
            item_project_no = item.get("project_no") or project_no
            item_project_name = item.get("project_name") or project_name
            row_values = [
                item.get("index", ""),
                item_project_no,
                item_project_name,
                item.get("material_name", ""),
                item.get("material_code", ""),
                item.get("qty", 0),
                item.get("price", 0),
                item.get("amount", 0),
                item.get("delivery_date", ""),
            ]
            x = start_x
            for i, v in enumerate(row_values):
                c.setFont(font_name, 10)
                c.drawString(x, y, safe_str(v))
                x += col_widths[i]
            y -= base_line_spacing
    else:
        row = 42
        while row <= ws.max_row and y > 80:
            values = [
                get_display_value_by_pos(row, 2),
                get_display_value_by_pos(row, 3),
                get_display_value_by_pos(row, 7),
                get_display_value_by_pos(row, 9),
                get_display_value_by_pos(row, 11),
                get_display_value_by_pos(row, 15),
                get_display_value_by_pos(row, 19),
                get_display_value_by_pos(row, 26),
                get_display_value_by_pos(row, 29),
            ]
            if isinstance(values[0], str) and "备注" in values[0]:
                break
            if all(v in [None, ""] for v in values):
                row += 1
                continue
            x = start_x
            for i, v in enumerate(values):
                c.setFont(font_name, 10)
                c.drawString(x, y, safe_str(v))
                x += col_widths[i]
            row_height = ws.row_dimensions[row].height or 15
            line_spacing = max(base_line_spacing, int(row_height + 4))
            y -= line_spacing
            row += 1

    c.save()


async def generate_contract_pdf(
    db: Session,
    inquiry_id: int,
    contract_template: ContractTemplate = None,
    template_file_path: str = None,
    buyer_company_name: str = None
) -> str:
    link = _resolve_deal_link(db, inquiry_id)
    active_template = contract_template or _get_active_contract_template(db)
    resolved_buyer_name = buyer_company_name or (active_template.default_buyer_name if active_template else None)
    contract_record = db.query(Contract).filter(Contract.inquiry_supplier_id == link.id).first()
    if not contract_record:
        contract_record = Contract(
            task_id=link.task_id,
            inquiry_supplier_id=link.id,
            status="pending",
            buyer_company_name=resolved_buyer_name or "俊朗电气有限公司",
        )
        db.add(contract_record)
        db.flush()
    payload = _collect_contract_payload(
        db,
        link,
        contract_record=contract_record,
        buyer_company_name=resolved_buyer_name
    )

    output_pdf = CONTRACT_DIR / f"合同_{inquiry_id}.pdf"
    resolved_template_file_path = template_file_path or (active_template.file_path if active_template else None)
    template_path = _resolve_template_path(resolved_template_file_path)
    temp_xlsx = _fill_template_to_temp_excel(payload, template_path=template_path)
    try:
        _export_excel_to_pdf(temp_xlsx, output_pdf, payload=payload)
    finally:
        if temp_xlsx and temp_xlsx.exists():
            try:
                temp_xlsx.unlink()
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"无法删除临时Excel文件 {temp_xlsx}: {e}")

    static_pdf_path = f"/static/contracts/{output_pdf.name}"
    if contract_record.pdf_path and contract_record.pdf_path != static_pdf_path:
        contract_record.history_versions = _append_history_version(
            contract_record.history_versions,
            contract_record.pdf_path
        )
    contract_record.pdf_path = static_pdf_path
    contract_record.status = "generated"
    contract_record.total_amount = payload.get("total_amount")
    if payload.get("buyer_name"):
        contract_record.buyer_company_name = payload.get("buyer_name")
    db.add(contract_record)
    db.commit()
    db.refresh(contract_record)
    return static_pdf_path


async def generate_contract_pdf_from_mock_data(mock_data: dict, output_filename: str = "test_result.pdf") -> str:
    payload = {
        "contract_no": mock_data.get("contract_no", f"TEST-{datetime.now().strftime('%Y%m%d%H%M%S')}"),
        "supplier_name": mock_data.get("supplier_name", ""),
        "buyer_name": mock_data.get("buyer_name", "需方"),
        "project_no": mock_data.get("project_no", ""),
        "task_title": mock_data.get("task_title", "测试合同"),
        "items": mock_data.get("items", []),
        "total_amount": float(mock_data.get("total_amount", 0)),
    }
    output_pdf = CONTRACT_DIR / output_filename
    temp_xlsx = _fill_template_to_temp_excel(payload)
    try:
        _export_excel_to_pdf(temp_xlsx, output_pdf, payload=payload)
    finally:
        if temp_xlsx and temp_xlsx.exists():
            try:
                temp_xlsx.unlink()
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"无法删除临时Excel文件 {temp_xlsx}: {e}")
    return str(output_pdf)
